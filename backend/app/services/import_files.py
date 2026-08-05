"""Reading a stock sheet: bytes in, validated rows out.

Deliberately free of database and HTTP concerns so the parsing rules can be
tested against real files without a session or a client. ``imports.py`` owns
persistence.

Three things here are answers to real failures rather than defensive padding:

* **Encoding.** ERP exports are frequently Windows-1252, not UTF-8 (Q4). A
  naive ``decode("utf-8")`` fails on the first rupee sign or accented name.
* **Delimiter.** Semicolon-separated CSV is what Excel writes under a European
  locale. Sniffed, not assumed.
* **A title row above the headers.** The plan names this as the exact reason
  one of the sample files failed, so the header row is *found* rather than
  assumed to be row 1.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from collections.abc import Callable
from dataclasses import dataclass, field, replace
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from app.core.errors import AppError
from app.models import COMPLAINT_COLUMNS, COUNT_COLUMNS, normalize_sku

# Canonical field -> header spellings seen in the wild. Compared after the same
# normalisation applied to the header cell, so "Item Code", "item_code" and
# "ITEM-CODE" all collapse to one entry.
FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "sku": (
        "sku",
        "skucode",
        "productsku",
        "itemsku",
        "productcode",
        "itemcode",
        "articlecode",
        "stylecode",
        "code",
    ),
    "product_name": (
        "productname",
        "name",
        "title",
        "itemname",
        "product",
        "description",
        "productdescription",
    ),
    # NOTE "totalqty" is deliberately *not* here. The sheet carries Quantity and
    # Total Qty as two different columns, so mapping both to one field would
    # silently discard one of them.
    "quantity": (
        "quantity",
        "qty",
        "stock",
        "stockqty",
        "stockquantity",
        "stockcount",
        "currentstock",
        "quantityonhand",
        "onhand",
        "closingstock",
        "available",
        "availableqty",
        "availablequantity",
        "inventory",
        "inventoryqty",
        "inventoryquantity",
    ),
    "price": ("price", "mrp", "rate", "sellingprice", "unitprice", "cost"),
    "category": ("category", "producttype", "type", "group", "productgroup"),
    # --- the raw complaint export's own columns (Format 2) -----------------
    #: One complaint per row, in the operator's words. Its presence is what
    #: tells the two formats apart.
    "reason": ("reason", "complaint", "issue", "issuetype", "complainttype", "reasontype"),
    #: Distinct order numbers become Total Orders. Without it every row counts
    #: as its own order, which is the best available reading.
    "order_no": (
        "orderno",
        "ordernumber",
        "orderid",
        "order",
        "invoiceno",
        "invoicenumber",
        "awb",
        "awbno",
    ),
    "employee": ("employee", "employeename", "staff", "agent", "handledby", "user"),
    "date": ("date", "orderdate", "complaintdate", "createdat", "timestamp"),
}

# **The SKU is the only required column.** It is the join key — a row without
# one cannot be reconciled against anything, and there is nothing to infer it
# from. Everything else is optional and defaults to zero, because a sheet that
# carries fewer columns is still worth importing: the alternative is refusing
# real data over a column the user does not have.
REQUIRED_FIELDS = ("sku",)

# The stock figure, in order of preference. Absent entirely is allowed; the
# quantity is then zero, or — in the raw complaint format — derived from the rows.
QUANTITY_FIELDS = ("total_qty", "quantity")

# Bounded so a malformed file cannot make the search run over a whole sheet.
MAX_HEADER_SCAN_ROWS = 10

_NON_ALNUM = re.compile(r"[^a-z0-9]+")
# The first number in the cell, so currency symbols and abbreviations around it
# are ignored rather than merged into the value.
_MONEY_NUMBER = re.compile(r"-?\d+(?:\.\d+)?")


class ImportFileError(AppError):
    """The file cannot be read at all. Nothing is written when this is raised."""

    code = "import_unreadable"
    status_code = 422
    message = "That file couldn't be read."
    next_step = "Check it opens in Excel, then upload it again."


class UnsupportedFileTypeError(ImportFileError):
    code = "unsupported_file_type"
    message = "That file type isn't supported."
    next_step = "Upload a .csv, .xlsx or .xls file."


class MissingHeadersError(ImportFileError):
    code = "missing_headers"
    message = "That sheet is missing a column StockSync Analytics needs."
    next_step = "Add the missing column, or rename yours to match, and upload again."


class EmptyFileError(ImportFileError):
    code = "empty_file"
    message = "That file has no rows in it."
    next_step = "Check you exported the right sheet, then upload it again."


def normalize_header(value: object) -> str:
    """Fold a header cell to its comparison form."""
    return _NON_ALNUM.sub("", str(value or "").strip().lower())


# The fixed part of the sheet: three count columns and ten complaint columns.
# Their aliases are *derived* from the headers in app.models.inventory rather
# than retyped, so adding a category there is the only edit needed.
SHEET_COLUMNS: tuple[tuple[str, str], ...] = COUNT_COLUMNS + COMPLAINT_COLUMNS

for _field, _header in SHEET_COLUMNS:
    FIELD_ALIASES[_field] = (normalize_header(_header),)

#: Spellings beyond the sheet's own headers. Merged in rather than replacing,
#: so the canonical header always matches and these are alternatives to it.
_EXTRA_ALIASES: dict[str, tuple[str, ...]] = {
    "total_count": ("count", "totalcount", "records", "lines"),
    "total_orders": ("orders", "ordercount", "totalorders", "nooforders", "orderqty"),
    # Deliberately *not* "qty" or "quantity": those belong to the `quantity`
    # field, and adding them here would make the two fields fight over one
    # column on a sheet that carries both. The exact header `Total Qty` always
    # wins over everything on this line — see `_match_columns` — and a sheet
    # whose only stock column is called Quantity still fills Total Qty through
    # the mirror in `_counts_for`.
    "total_qty": ("units", "totalquantity", "pcs"),
}
for _field, _extra in _EXTRA_ALIASES.items():
    FIELD_ALIASES[_field] = tuple(dict.fromkeys(FIELD_ALIASES[_field] + _extra))

#: What each field is *called*, as opposed to what it also answers to. The first
#: alias is the canonical spelling by construction — the sheet's own header for
#: every count and complaint column, and the obvious name for the rest — and
#: ``_match_columns`` gives it priority over every nickname below it.
CANONICAL_HEADERS: dict[str, str] = {
    field: aliases[0] for field, aliases in FIELD_ALIASES.items() if aliases
}

#: A free-text reason, folded, mapped to the complaint column it belongs in.
#:
#: **Ordered most specific first and matched in that order**, because the
#: general terms are substrings of the specific ones: "missing part" contains
#: "missing", and "damage partial" contains "damage". Testing the general case
#: first would file every missing part under Missing.
#:
#: Ordering alone is not enough for the partial/complete pairs, though — see
#: ``PARTIAL_MARKERS`` and ``_QUALIFIED_FAMILIES`` below for the rest of it.
REASON_COLUMNS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("missing_part", ("missingpart", "missingparts", "partmissing", "partsmissing", "incomplete")),
    (
        "item_mismatch_wrong_item",
        (
            "wrongitem",
            "wrongitemdelivered",
            "itemmismatch",
            "mismatch",
            "wrongproduct",
            "wrongsku",
            "differentitem",
        ),
    ),
    ("order_wrong_parcel", ("wrongparcel", "wrongorder", "wrongpackage", "parcelmismatch")),
    (
        "electronics_nonworking_partial",
        ("nonworkingpartial", "partialnonworking", "partiallyworking", "notworkingpartial"),
    ),
    (
        "electronics_nonworking_complete",
        (
            "nonworking",
            "notworking",
            "nonworkingcomplete",
            "notworkingcomplete",
            "deadonarrival",
            "doa",
            "notswitchingon",
        ),
    ),
    ("item_damage_partial", ("damagepartial", "partialdamage", "partiallydamaged")),
    (
        "item_damage_complete",
        ("damage", "damaged", "damagecomplete", "completedamage", "fullydamaged", "broken"),
    ),
    ("item_defect_partial", ("defectpartial", "partialdefect", "partiallydefective")),
    (
        "item_defect_complete",
        (
            "defect",
            "defective",
            "defectcomplete",
            "completedefect",
            "faulty",
            "manufacturingdefect",
        ),
    ),
    ("missing", ("missing", "itemmissing", "notreceived", "lost", "shortage")),
)


#: Words that say "some of it", however the operator spelled them.
#:
#: Checked as substrings of the folded reason, so "partially", "part-damaged"
#: and "partly working" all register. ``incomplete`` is deliberately absent: it
#: already belongs to ``missing_part`` as a whole word and contains "complete",
#: which would make it argue with itself.
PARTIAL_MARKERS: tuple[str, ...] = ("partial", "partly", "somewhat")

#: Words that say "all of it". Only consulted to break a tie the partial
#: markers did not already claim.
COMPLETE_MARKERS: tuple[str, ...] = ("complete", "fully", "entirely", "totally", "whole")

#: The three families that come in a partial/complete pair, keyed by the terms
#: that identify the *family* regardless of degree.
#:
#: This exists because ordering could not fix the pairs on its own. The partial
#: aliases above are contiguous compounds — ``damagepartial``, ``partialdamage``
#: — so a reason like "damaged in transit - partial" folds to
#: ``damagedintransitpartial``, matches none of them, and then falls through to
#: ``item_damage_complete``, whose alias ``damage`` *is* a substring of it. A
#: partial return was recorded as a total loss, and the same held for
#: "defective - partial" and "item damaged partially".
#:
#: Family terms are matched most specific first within each tuple for the same
#: substring reason as everywhere else here.
_QUALIFIED_FAMILIES: tuple[tuple[tuple[str, ...], str, str], ...] = (
    (
        ("nonworking", "notworking", "notswitchingon", "deadonarrival", "doa"),
        "electronics_nonworking_partial",
        "electronics_nonworking_complete",
    ),
    (
        ("damage", "damaged", "broken"),
        "item_damage_partial",
        "item_damage_complete",
    ),
    (
        ("defect", "defective", "faulty"),
        "item_defect_partial",
        "item_defect_complete",
    ),
)


def _degree(folded: str) -> str | None:
    """ "partial", "complete", or None when the reason does not say."""
    if any(marker in folded for marker in PARTIAL_MARKERS):
        return "partial"
    if any(marker in folded for marker in COMPLETE_MARKERS):
        return "complete"
    return None


def complaint_for(reason: object) -> str | None:
    """The complaint column a reason belongs in, or None if it is unrecognised.

    Three passes, in this order:

    1. **Exact fold.** "damagepartial" is unambiguous; nothing should override it.
    2. **Family plus degree.** If the reason names one of the three paired
       families *and* says partial or complete anywhere in the text, that
       decides it. This runs before the general substring pass so a "partial"
       sitting at the far end of a sentence still wins over a bare family term
       in the middle of it.
    3. **General substring**, most specific alias first, for the unpaired
       columns and for family terms with no degree stated. A bare "Damage"
       still means complete damage, which is the reading operators intend.

    An unrecognised reason is not an error: the row still counts towards Total
    Count, Total Orders and Total Qty. Only the complaint breakdown loses it, and
    the parse result names the reasons it could not place so the mapping can be
    extended rather than the data quietly under-reporting.
    """
    folded = normalize_header(reason)
    if not folded:
        return None

    for field_name, aliases in REASON_COLUMNS:
        if folded in aliases:
            return field_name

    degree = _degree(folded)
    if degree is not None:
        for terms, partial_field, complete_field in _QUALIFIED_FAMILIES:
            if any(term in folded for term in terms):
                return partial_field if degree == "partial" else complete_field

    for field_name, aliases in REASON_COLUMNS:
        if any(alias in folded for alias in aliases):
            return field_name
    return None


#: How to spell a field when telling someone their sheet is missing it. The
#: aliases are folded ("totalqty"), which is right for matching and useless in a
#: message, so the required fields carry a human form alongside.
DISPLAY_HEADERS: dict[str, str] = {"sku": "SKU", "total_qty": "Total Qty.", "quantity": "Quantity"}


@dataclass(frozen=True)
class ParsedRow:
    """One accepted row, already coerced to storage types."""

    row_number: int
    sku: str
    sku_normalized: str
    product_name: str
    quantity: int
    price_paise: int | None
    category: str | None
    #: The fixed count and complaint columns, keyed by model attribute. A dict
    #: rather than thirteen fields: every consumer treats them as a group, and a
    #: new category should not mean touching this dataclass.
    counts: dict[str, int] = field(default_factory=dict)
    #: Complaints by the day they happened, for the rows that carried a date:
    #: ``{date: {category: count}}``. Empty for the aggregated format, which has
    #: no dates at all — see ``undated_complaints``.
    dated_counts: dict[date, dict[str, int]] = field(default_factory=dict)
    #: Complaints this row could not place on a day. Everything from an
    #: aggregated sheet, plus any complaint row whose date cell was blank or
    #: unreadable. Carried rather than dropped so the totals still add up and
    #: the UI can say how much of the picture a date filter cannot see.
    undated_complaints: int = 0


@dataclass(frozen=True)
class RejectedRow:
    """One row that could not be stored, with the reason in the user's terms."""

    row_number: int
    reason: str
    detail: str


@dataclass
class DuplicateGroup:
    """A SKU that appeared more than once in the same file."""

    sku: str
    rows: list[int] = field(default_factory=list)
    merged_quantity: int = 0


#: What the sheet turned out to be. "aggregated" is one row per SKU with the
#: counts already totalled; "complaints" is one row per complaint, which this
#: module groups by SKU itself.
log = logging.getLogger(__name__)

SheetFormat = str

AGGREGATED: SheetFormat = "aggregated"
COMPLAINTS: SheetFormat = "complaints"

#: A Date column was recognised and then not used, because the sheet has no
#: Reason column and so was read as one row per SKU.
DATE_COLUMN_IGNORED = "date_column_ignored"
#: The same, and rows for the same SKU were merged — which for a per-day file
#: means quantities were summed across days. The more serious of the two.
DATE_COLUMN_IGNORED_WITH_DUPLICATES = "date_column_ignored_with_duplicates"


@dataclass
class ParseResult:
    rows: list[ParsedRow]
    rejected: list[RejectedRow]
    duplicates: list[DuplicateGroup]
    header_row_number: int
    detected_columns: dict[str, str]
    rows_read: int
    #: Which of the two shapes the sheet was read as.
    sheet_format: SheetFormat = AGGREGATED
    #: Reasons the mapping could not place, with how many rows carried each.
    #: Surfaced rather than swallowed: those rows counted towards the totals but
    #: not towards any complaint column, and silence would look like clean data.
    unmapped_reasons: dict[str, int] = field(default_factory=dict)
    #: Things the reader noticed that the user needs to know but that are not
    #: failures. Stable codes; the client words them, because display text on the
    #: server is text that goes stale where nobody looks.
    warnings: list[str] = field(default_factory=list)

    @property
    def rows_merged(self) -> int:
        """Row occurrences absorbed into an earlier row (not the first of each)."""
        return sum(len(group.rows) - 1 for group in self.duplicates)


# ---------------------------------------------------------------------------
# decoding
# ---------------------------------------------------------------------------


def _decode(raw: bytes) -> str:
    """Text from bytes, trying the encodings ERP exports actually use.

    utf-8-sig first because it also strips the BOM Excel writes, which would
    otherwise become part of the first header name and break matching.
    """
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    # latin-1 maps every byte, so reaching here means the input was not text.
    raise ImportFileError


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        # Sniffer fails on a single-column file, where any delimiter is correct.
        return ","


# ---------------------------------------------------------------------------
# header discovery
# ---------------------------------------------------------------------------


def _match_columns(header: list[Any]) -> dict[str, int]:
    """Map canonical field -> column index for one candidate header row.

    **Two passes, and the order matters.** Every field first tries to claim the
    column named exactly what that field is called; only then do aliases get a
    turn, and only over columns nobody has claimed.

    One pass was wrong in a way that looked fine. Each field took the *first*
    column matching *any* of its aliases, so a sheet headed
    ``SKU | Units | Total Qty.`` bound ``total_qty`` to ``Units`` — "units" is
    an alias — and the column actually named ``Total Qty.`` was never read. The
    imported figure depended on which column came first, not on what it was
    called, and nothing reported it. An exact name must beat a nickname.

    Claiming also stops two fields resolving to one column, which the old code
    allowed silently.
    """
    normalized = [normalize_header(cell) for cell in header]
    found: dict[str, int] = {}
    claimed: set[int] = set()

    def claim(canonical: str, index: int) -> None:
        found[canonical] = index
        claimed.add(index)

    for canonical, exact in CANONICAL_HEADERS.items():
        for index, cell in enumerate(normalized):
            if index not in claimed and cell and cell == exact:
                claim(canonical, index)
                break

    for canonical, aliases in FIELD_ALIASES.items():
        if canonical in found:
            continue
        for index, cell in enumerate(normalized):
            if index not in claimed and cell and cell in aliases:
                claim(canonical, index)
                break

    return found


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    """Locate the header row and its column mapping.

    Scans the first few rows rather than trusting row 1, because a report
    exported with a title line above the table is a real and common shape. The
    first row that resolves every required field wins.
    """
    best_index = 0
    best_match: dict[str, int] = {}
    for index, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
        match = _match_columns(row)
        if _resolves(match):
            return index, match
        if len(match) > len(best_match):
            best_index, best_match = index, match

    # Nothing resolved. Report against the closest candidate so the message
    # names the columns the user actually has.
    missing = [f for f in REQUIRED_FIELDS if f not in best_match]
    present = [str(c).strip() for c in rows[best_index] if str(c or "").strip()]
    wanted = [DISPLAY_HEADERS.get(f, f) for f in missing]
    raise MissingHeadersError(
        # Specific, not generic. The old message — "That sheet is missing a
        # column StockSync Analytics needs" — was true of every rejection and
        # useful for none, and it is the string Import History keeps, so a
        # failure from last week could not be diagnosed at all.
        _missing_headers_message(wanted, present),
        detail={
            "missing": missing,
            "expected_columns": wanted,
            "found_columns": present[:40],
            "accepted_names": {f: list(FIELD_ALIASES[f]) for f in missing},
            "header_row_number": best_index + 1,
        },
    )


def _missing_headers_message(wanted: list[str], found: list[str]) -> str:
    """Name what is missing and what was there instead, in one sentence.

    Both halves matter. "Missing Total Qty." alone leaves the user guessing
    whether the file was read at all; listing what *was* found is usually enough
    to see the real problem — the wrong sheet, a title row, an export with the
    columns renamed.
    """
    needed = wanted[0] if len(wanted) == 1 else " and ".join((", ".join(wanted[:-1]), wanted[-1]))
    plural = "column" if len(wanted) == 1 else "columns"
    # `found` is never empty: blank rows are dropped before the scan, so every
    # candidate row has at least one non-empty cell, and a wholly blank file
    # raises EmptyFileError before reaching here.
    shown = ", ".join(found[:8])
    more = f" and {len(found) - 8} more" if len(found) > 8 else ""
    return f"That sheet has no {needed} {plural}. Found: {shown}{more}."


def _resolves(match: dict[str, int]) -> bool:
    """A usable header row. A SKU is enough; everything else is optional."""
    return all(f in match for f in REQUIRED_FIELDS)


# ---------------------------------------------------------------------------
# value coercion
# ---------------------------------------------------------------------------


#: The date spellings an ERP export actually writes. ISO first because it is
#: unambiguous; the day-first forms next because this is an Indian business and
#: 03/08/2026 means 3 August here. A month-first reading is deliberately absent:
#: guessing between the two silently moves complaints between months.
_DATE_FORMATS: tuple[str, ...] = (
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%d.%m.%Y",
    "%d-%m-%y",
    "%d/%m/%y",
    "%Y/%m/%d",
    "%d-%b-%Y",
    "%d %b %Y",
    "%d-%B-%Y",
    "%d %B %Y",
)


def parse_complaint_date(value: object) -> date | None:
    """The date on a complaint row, or None when there isn't a usable one.

    None is not an error. A complaint with no date still counts towards every
    total the sheet reports; it simply cannot be placed in a window, and
    ``imports`` records it as undated rather than assigning it a day that would
    then be filtered on.

    Excel hands us a real ``datetime`` for a date-formatted cell, so that case
    is taken directly rather than round-tripped through a string.
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _clean(value)
    if not text:
        return None

    # The whole cell first. Dropping everything after the first space would be
    # simpler and would break "14 July 2026", where the spaces are part of the
    # date rather than a separator before a time.
    candidates = [text]
    # Then the leading date of a timestamp: "2026-01-31 14:03:00" and
    # "2026-01-31T14:03:00" both carry the day first.
    head = text.replace("T", " ").split(" ")[0]
    if head != text:
        candidates.append(head)

    for candidate in candidates:
        for fmt in _DATE_FORMATS:
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
    return None


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


#: A line break inside a SKU cell, in any of the spellings a sheet can produce.
_SKU_LINE_BREAK = re.compile(r"[\n\r\v\f  ]")
#: Two or more spaces, or a tab, inside a SKU that is otherwise one value.
_SKU_RUN_OF_SPACE = re.compile(r"[ \t ]{1,}")


class SkuIsTwoValuesError(ValueError):
    """A SKU cell that holds more than one SKU.

    Raised rather than repaired: the two halves cannot be told apart with any
    confidence, and inventing a single value from them is the failure this
    guards against.
    """


def parse_sku(value: object) -> str:
    """The SKU as it should be stored, or a refusal.

    Two different problems arrive in this cell and they need opposite answers:

    * **A line break** means two rows were pasted into one — a real workspace
      had ``'14124_smart_floor_cleaning_robot_1pc\\n4710_airplane_launcher_gun_toy'``,
      which is two products. Collapsing it would mint a SKU that matches
      neither, silently losing both, so the row is refused and reported by
      number like any other bad row. :class:`SkuIsTwoValuesError`.
    * **A run of spaces or a tab** is a formatting artefact of one SKU —
      ``'DD 1001'`` typed with two spaces is still ``DD 1001``. Those collapse
      to a single space, which is tidying rather than merging.

    Leading and trailing whitespace is stripped as it always was. Everything
    else about the value is left alone: the SKU is the user's identifier, not
    ours to rewrite.
    """
    text = _clean(value)
    if not text:
        return ""
    if _SKU_LINE_BREAK.search(text):
        raise SkuIsTwoValuesError(text)
    return _SKU_RUN_OF_SPACE.sub(" ", text)


def _two_values_message(exc: SkuIsTwoValuesError) -> str:
    """Name both halves, so the row can be found and split in the source sheet."""
    halves = [part.strip() for part in _SKU_LINE_BREAK.split(str(exc)) if part.strip()]
    named = " and ".join(repr(half) for half in halves[:2])
    return (
        f"This SKU cell holds more than one SKU ({named}). "
        "Split it across two rows and upload again."
    )


def parse_count(value: object) -> int:
    """A sheet count column, coerced to a non-negative whole number.

    Unlike :func:`parse_quantity` this never fails the row. These columns are
    tallies: blank means zero, and a stray value in one complaint cell is not a
    reason to discard the SKU's stock figure.
    """
    parsed = parse_quantity(value)
    return max(0, parsed) if parsed is not None else 0


def parse_quantity(value: object) -> int | None:
    """Whole units, or None if the cell is not a usable quantity.

    Accepts "1,240" and "12.0": spreadsheet cells arrive as floats even when
    the column only ever holds integers, and thousands separators survive a CSV
    export. A fractional quantity is rejected rather than rounded — silently
    turning 2.5 into 2 or 3 is a stock error nobody would catch.
    """
    text = _clean(value).replace(",", "").replace(" ", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if number != int(number):
        return None
    return int(number)


def parse_price_paise(value: object) -> int | None:
    """Rupees to paise. None when absent or unparseable.

    The number is *matched* rather than the non-numeric characters deleted.
    Deleting them turns "Rs. 45" into ".45" — the full stop in the currency
    abbreviation survives and the price silently becomes 45 paise.

    Rounded at the last step because 12.99 is not representable in binary
    floating point and int(12.99 * 100) is 1298.
    """
    text = _clean(value).replace(",", "")
    if not text:
        return None
    found = _MONEY_NUMBER.search(text)
    if found is None:
        return None
    try:
        return round(float(found.group()) * 100)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# the parse
# ---------------------------------------------------------------------------


def _rows_from_csv(raw: bytes) -> list[list[Any]]:
    text = _decode(raw)
    if not text.strip():
        raise EmptyFileError
    delimiter = _sniff_delimiter(text[:8192])
    return [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _rows_from_xlsx(raw: bytes) -> list[list[Any]]:
    try:
        # read_only streams rather than building the whole object graph;
        # data_only takes the cached value of a formula rather than "=A1*2".
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a wide range on malformed input
        raise ImportFileError from exc

    try:
        sheet = workbook.worksheets[0]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def parse_inventory_file(filename: str, raw: bytes) -> ParseResult:
    """Read a stock sheet into rows ready to store.

    Duplicate SKUs inside one file are merged by summing quantity, per design
    doc §8.6. Q5 is open on whether that is right for every column — a SKU
    listed twice at two prices sums the quantity correctly but has to pick a
    price, and this takes the first non-empty one.
    """
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix == "csv":
        grid = _rows_from_csv(raw)
    elif suffix in ("xlsx", "xlsm"):
        grid = _rows_from_xlsx(raw)
    elif suffix == "xls":
        # openpyxl reads the OOXML format only. Saying so beats a stack trace
        # about a corrupt zip archive.
        raise UnsupportedFileTypeError(
            "That looks like the older .xls format, which StockSync Analytics can't read.",
            next_step='Open it in Excel and use "Save As" to save a .xlsx, then upload that.',
        )
    else:
        raise UnsupportedFileTypeError

    grid = [row for row in grid if any(_clean(cell) for cell in row)]
    if not grid:
        raise EmptyFileError

    header_index, columns = _find_header(grid)
    body = grid[header_index + 1 :]
    if not body:
        raise EmptyFileError

    detected = {name: str(grid[header_index][i]) for name, i in columns.items()}

    # A `reason` column means one row per complaint rather than one row per SKU.
    # That is the only reliable signal: the raw export shares SKU with the
    # aggregated one and carries none of its count columns, while an aggregated
    # sheet never has a free-text reason.
    if "reason" in columns:
        return _parse_complaints(body, columns, header_index, detected)

    parsed = _parse_aggregated(body, columns, header_index, detected)

    # **A Date column that was matched and then not used must say so.**
    #
    # `date` is a column this reader recognises, so it lands in
    # `detected_columns` and the import screen prints "matched Date → date" —
    # which reads as confirmation the dates were used. Without a `reason` column
    # the aggregated reader never looks at them, so they were silently dropped.
    #
    # Worse, and the reason this is a warning rather than a note: a file carrying
    # one row per SKU *per day* has no reason column either, so its per-day rows
    # are merged as duplicates and their quantities summed. Ten units on Monday
    # and ten on Tuesday become twenty in stock.
    if "date" in columns:
        parsed.warnings.append(
            DATE_COLUMN_IGNORED_WITH_DUPLICATES if parsed.duplicates else DATE_COLUMN_IGNORED
        )
        log.warning(
            "import: a Date column (%r) was found but the sheet has no Reason column, "
            "so it was read as one row per SKU and the dates were not used "
            "(%s duplicate group(s) merged)",
            detected.get("date"),
            len(parsed.duplicates),
        )
    return parsed


def _counts_for(
    cell: Callable[[list[Any], str], object],
    row: list[Any],
    columns: dict[str, int],
    quantity: int,
) -> dict[str, int]:
    """The thirteen count columns for one aggregated row.

    Absent or unreadable counts as zero: a blank complaint cell means no
    complaints, not unknown, and rejecting the row over it would throw away good
    stock data.

    ``total_qty`` mirrors the resolved quantity when the sheet has no Total Qty
    column of its own — a sheet whose stock column is called `Qty` or `Quantity`
    means that as its total, and leaving the field at zero would report a stocked
    SKU as holding nothing.
    """
    counts = {field_name: parse_count(cell(row, field_name)) for field_name, _ in SHEET_COLUMNS}
    if "total_qty" not in columns:
        counts["total_qty"] = quantity
    return counts


def _parse_aggregated(
    body: list[list[Any]],
    columns: dict[str, int],
    header_index: int,
    detected: dict[str, str],
) -> ParseResult:
    """Format 1: one row per SKU, counts already totalled. Taken as given."""

    def cell(row: list[Any], name: str) -> object:
        index = columns.get(name)
        if index is None or index >= len(row):
            return None
        return row[index]

    accepted: dict[str, ParsedRow] = {}
    order: list[str] = []
    rejected: list[RejectedRow] = []
    duplicates: dict[str, DuplicateGroup] = {}

    for offset, row in enumerate(body):
        # +2 converts to the 1-based line the user sees in Excel: +1 for the
        # header row itself, +1 because spreadsheet rows start at 1.
        row_number = header_index + offset + 2

        try:
            sku = parse_sku(cell(row, "sku"))
        except SkuIsTwoValuesError as exc:
            rejected.append(RejectedRow(row_number, "invalid_sku", _two_values_message(exc)))
            continue
        if not sku:
            rejected.append(
                RejectedRow(row_number, "missing_sku", "No SKU in this row, so nothing to match.")
            )
            continue

        # No quantity column at all is allowed — the sheet is still worth
        # importing for its SKUs and complaint counts. Only a column that is
        # present and unreadable rejects the row.
        source = next((f for f in QUANTITY_FIELDS if f in columns), None)
        quantity = 0
        if source is not None:
            parsed_quantity = parse_quantity(cell(row, source))
            label = "Total Qty." if source == "total_qty" else "Quantity"
            if parsed_quantity is None:
                rejected.append(
                    RejectedRow(
                        row_number,
                        "bad_quantity",
                        f"{label} {_clean(cell(row, source))!r} isn't a whole number.",
                    )
                )
                continue
            if parsed_quantity < 0:
                rejected.append(
                    RejectedRow(
                        row_number, "bad_quantity", f"{label} {parsed_quantity} is negative."
                    )
                )
                continue
            quantity = parsed_quantity

        key = normalize_sku(sku)
        if not key:
            rejected.append(
                RejectedRow(
                    row_number, "missing_sku", f"SKU {sku!r} has no letters or digits in it."
                )
            )
            continue

        parsed = ParsedRow(
            row_number=row_number,
            sku=sku,
            sku_normalized=key,
            product_name=_clean(cell(row, "product_name")),
            quantity=quantity,
            price_paise=parse_price_paise(cell(row, "price")),
            category=_clean(cell(row, "category")) or None,
            # Absent or unreadable counts as zero: a blank complaint cell means
            # no complaints, not unknown, and rejecting the row over it would
            # throw away good stock data.
            counts=_counts_for(cell, row, columns, quantity),
        )
        # An aggregated sheet is one row per SKU with running totals on it. There
        # is no date to be had, so every complaint on it is undated.
        parsed = replace(
            parsed,
            undated_complaints=sum(parsed.counts[f] for f, _ in COMPLAINT_COLUMNS),
        )

        existing = accepted.get(key)
        if existing is None:
            accepted[key] = parsed
            order.append(key)
            continue

        # Same SKU twice in one file: sum the quantity, keep the first row's
        # descriptive fields, and record the group so the summary can show it.
        group = duplicates.setdefault(
            key, DuplicateGroup(sku=existing.sku, rows=[existing.row_number])
        )
        group.rows.append(row_number)
        merged_quantity = existing.quantity + parsed.quantity
        group.merged_quantity = merged_quantity
        accepted[key] = ParsedRow(
            row_number=existing.row_number,
            sku=existing.sku,
            sku_normalized=key,
            product_name=existing.product_name or parsed.product_name,
            quantity=merged_quantity,
            price_paise=existing.price_paise
            if existing.price_paise is not None
            else parsed.price_paise,
            category=existing.category or parsed.category,
            # Summed, like quantity. Two warehouse rows for one SKU carry two
            # halves of the same complaint tally.
            counts={
                field: existing.counts.get(field, 0) + parsed.counts.get(field, 0)
                for field, _header in SHEET_COLUMNS
            },
            undated_complaints=existing.undated_complaints + parsed.undated_complaints,
        )

    return ParseResult(
        rows=[accepted[key] for key in order],
        rejected=rejected,
        duplicates=list(duplicates.values()),
        header_row_number=header_index + 1,
        detected_columns=detected,
        rows_read=len(body),
        sheet_format=AGGREGATED,
    )


# ---------------------------------------------------------------------------
# Format 2 — one row per complaint
# ---------------------------------------------------------------------------


@dataclass
class _Tally:
    """One SKU's running totals while the complaint rows are walked."""

    sku: str
    first_row: int
    rows: list[int] = field(default_factory=list)
    #: Distinct order numbers. A set, because the same order can appear on
    #: several rows — two faulty items in one parcel is one order, two rows.
    orders: set[str] = field(default_factory=set)
    #: Rows with no order number. Counted individually: nothing distinguishes
    #: them, so treating them as one order would under-count, and the honest
    #: reading of a blank is "an order we cannot identify".
    orderless: int = 0
    quantity: int = 0
    counts: dict[str, int] = field(default_factory=dict)
    #: The same counts, split by the day each complaint carried.
    by_date: dict[date, dict[str, int]] = field(default_factory=dict)
    #: Complaints whose date cell was blank or unreadable. They still count
    #: towards every total; they just cannot be placed in a window.
    undated: int = 0


def _parse_complaints(
    body: list[list[Any]],
    columns: dict[str, int],
    header_index: int,
    detected: dict[str, str],
) -> ParseResult:
    """Format 2: one row per complaint, grouped into one row per SKU.

    ``Date | Order No | SKU Code | Reason | Employee`` and anything like it. The
    three count columns are *derived* here rather than read:

    * **Total Count** — rows for this SKU. One complaint, one row.
    * **Total Orders** — distinct order numbers, so two items from one parcel
      count once. No order column at all means one order per row.
    * **Total Qty** — the quantity column if the export has one, otherwise one
      per row, as the requirement specifies.

    The reason becomes a complaint column via ``complaint_for``. A reason that
    cannot be placed still counts towards all three totals; only the breakdown
    misses it, and it is reported in ``unmapped_reasons``.

    **The date is kept.** Each complaint is bucketed on the day it happened, so
    Complaint Rate % can divide complaints in a window by sales in the same
    window. A row whose date is blank or unreadable is counted as undated rather
    than dropped or guessed at — it still reaches every total, and the UI says
    how many there are so a filtered view never looks complete when it isn't.
    """

    def cell(row: list[Any], name: str) -> object:
        index = columns.get(name)
        if index is None or index >= len(row):
            return None
        return row[index]

    quantity_column = next((f for f in QUANTITY_FIELDS if f in columns), None)
    has_orders = "order_no" in columns

    tallies: dict[str, _Tally] = {}
    order: list[str] = []
    rejected: list[RejectedRow] = []
    unmapped: dict[str, int] = {}

    for offset, row in enumerate(body):
        row_number = header_index + offset + 2

        try:
            sku = parse_sku(cell(row, "sku"))
        except SkuIsTwoValuesError as exc:
            rejected.append(RejectedRow(row_number, "invalid_sku", _two_values_message(exc)))
            continue
        key = normalize_sku(sku)
        if not sku or not key:
            rejected.append(
                RejectedRow(row_number, "missing_sku", "No SKU in this row, so nothing to match.")
            )
            continue

        tally = tallies.get(key)
        if tally is None:
            tally = _Tally(
                sku=sku,
                first_row=row_number,
                counts={field_name: 0 for field_name, _ in SHEET_COLUMNS},
            )
            tallies[key] = tally
            order.append(key)

        tally.rows.append(row_number)

        order_no = _clean(cell(row, "order_no")) if has_orders else ""
        if order_no:
            tally.orders.add(order_no.lower())
        else:
            tally.orderless += 1

        # One per row unless the export counts them for us. A quantity that will
        # not parse falls back to 1 rather than rejecting a complaint that did
        # happen — the reason is the point of the row, not the number.
        units = 1
        if quantity_column is not None:
            parsed_units = parse_quantity(cell(row, quantity_column))
            if parsed_units is not None and parsed_units >= 0:
                units = parsed_units
        tally.quantity += units

        raw_reason = _clean(cell(row, "reason"))
        field_name = complaint_for(raw_reason)
        if field_name is not None:
            tally.counts[field_name] += 1
            # Bucketed on its own day. Only mapped reasons land here, so the
            # dated breakdown always sums to the dated total — an unmapped
            # reason is not a complaint in any category, dated or not.
            day = parse_complaint_date(cell(row, "date"))
            if day is None:
                tally.undated += 1
            else:
                bucket = tally.by_date.setdefault(day, {name: 0 for name, _ in COMPLAINT_COLUMNS})
                bucket[field_name] += 1
        elif raw_reason:
            unmapped[raw_reason] = unmapped.get(raw_reason, 0) + 1

    rows: list[ParsedRow] = []
    duplicates: list[DuplicateGroup] = []
    for key in order:
        tally = tallies[key]
        total_orders = len(tally.orders) + tally.orderless
        counts = dict(tally.counts)
        counts["total_count"] = len(tally.rows)
        counts["total_orders"] = total_orders
        counts["total_qty"] = tally.quantity

        rows.append(
            ParsedRow(
                row_number=tally.first_row,
                sku=tally.sku,
                sku_normalized=key,
                product_name="",
                quantity=tally.quantity,
                price_paise=None,
                category=None,
                counts=counts,
                dated_counts={day: dict(by_cat) for day, by_cat in tally.by_date.items()},
                undated_complaints=tally.undated,
            )
        )
        # Every SKU with more than one complaint row is a group. Reported through
        # the same channel as a duplicate in an aggregated sheet, because it is
        # the same thing from the reader's side: several source rows became one.
        if len(tally.rows) > 1:
            duplicates.append(
                DuplicateGroup(sku=tally.sku, rows=tally.rows, merged_quantity=tally.quantity)
            )

    return ParseResult(
        rows=rows,
        rejected=rejected,
        duplicates=duplicates,
        header_row_number=header_index + 1,
        detected_columns=detected,
        rows_read=len(body),
        sheet_format=COMPLAINTS,
        unmapped_reasons=dict(sorted(unmapped.items(), key=lambda kv: (-kv[1], kv[0]))),
    )
