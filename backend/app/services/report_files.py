"""Rendering a report table as CSV, XLSX or PDF — design doc §12.1.

Each writer takes the same :class:`ReportTable` and returns bytes, so the three
formats cannot drift apart in content — only in presentation.

**The PDF is written here rather than with a library.** The plan's §6 rule is
to ask before adding a dependency, and PDF is the only one of the three
formats Python cannot already produce: CSV is in the standard library and
openpyxl is already a dependency for reading .xlsx uploads. A table of text in
one of the base-14 fonts is a small, completely specified corner of PDF 1.4 —
those fonts are the ones every viewer is required to provide, so nothing has to
be embedded, and the result is a few hundred lines with no supply chain
attached. If richer output is ever wanted (charts, images, styled headers),
this is the file to replace with reportlab.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Sequence
from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.report_tables import ReportTable

CONTENT_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def render(table: ReportTable, fmt: str) -> bytes:
    if fmt == "csv":
        return to_csv(table)
    if fmt == "xlsx":
        return to_xlsx(table)
    if fmt == "pdf":
        return to_pdf(table)
    raise ValueError(f"unsupported report format {fmt!r}")


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


#: Leading characters that make a spreadsheet treat a cell as a formula rather
#: than as text. A tab or carriage return counts because Excel strips leading
#: whitespace before deciding, so " =cmd" is still a formula.
FORMULA_LEADERS = ("=", "+", "-", "@", "\t", "\r")


def is_formula_like(value: str) -> bool:
    """True when a spreadsheet would read this text as a formula."""
    return value.lstrip("\t\r ").startswith(FORMULA_LEADERS)


def neutralize(value: str) -> str:
    """Defuse a value a spreadsheet would otherwise execute.

    Product names and SKUs come from an uploaded sheet or from Shopify, so they
    are attacker-influenced text. ``=cmd|'/c calc'!A1`` in a product name is a
    working code-execution payload in Excel via DDE, and these exports exist to
    be emailed — so the person it runs on is often not the person who imported
    it. Prefixing with an apostrophe is the standard mitigation: Excel and
    LibreOffice both read the rest as literal text.

    CSV has no type system, so the guard has to live in the text itself. The
    XLSX writer does not use this — it can say "this cell is a string" properly.
    """
    return f"'{value}" if value and is_formula_like(value) else value


def to_csv(table: ReportTable) -> bytes:
    """UTF-8 with a BOM and CRLF line endings.

    Both are for Excel's benefit: without the BOM it reads the file in the
    system codepage and mangles the ₹ sign and any non-ASCII product name, and
    RFC 4180 specifies CRLF. Anything else reading the file copes with both.
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow([column.header for column in table.columns])
    writer.writerows([neutralize(value) for value in row] for row in table.rows)
    return buffer.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="EEF0F0")
_MAX_COLUMN_WIDTH = 52


def to_xlsx(table: ReportTable) -> bytes:
    """A real spreadsheet: frozen header, filters, and numbers stored as numbers."""
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:  # pragma: no cover - a new Workbook always has one
        sheet = workbook.create_sheet()
    # Excel rejects the characters below in a sheet name, and truncates at 31.
    sheet.title = table.title[:31].translate({ord(c): None for c in r"[]:*?/\\"})

    sheet.append([column.header for column in table.columns])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL

    widths = [len(column.header) for column in table.columns]
    formula_cells: list[tuple[int, int]] = []

    for row_index, row in enumerate(table.rows, start=2):
        # A figure written as text sorts and sums wrongly, and Excel flags every
        # cell with a green triangle. Right-aligned columns are the numeric ones.
        values: list[float | int | str] = []
        for column_index, (column, value) in enumerate(
            zip(table.columns, row, strict=False), start=1
        ):
            if column_index <= len(widths):
                widths[column_index - 1] = max(widths[column_index - 1], len(value))
            if column.align == "right":
                values.append(_as_number(value))
                continue
            values.append(value)
            if is_formula_like(value):
                formula_cells.append((row_index, column_index))
        sheet.append(values)

    # Anything that looks like a formula is pinned to "this is a string". Unlike
    # the CSV guard this leaves the value byte-for-byte intact — the cell simply
    # stops being a formula cell, so Excel has nothing to evaluate. openpyxl
    # infers ``data_type='f'`` from a leading '=' at assignment, which is what
    # made an injected product name a live formula in the workbook.
    for row_index, column_index in formula_cells:
        cell = sheet.cell(row=row_index, column=column_index)
        cell.data_type = "s"
        cell.quotePrefix = True

    # One Alignment instance for every right-aligned cell, not one each. Building
    # them per cell cost ~2.5 s of a 12.6 s export at 13,000 rows, and the
    # written file is identical either way.
    right = Alignment(horizontal="right")
    for index, width in enumerate(widths, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = min(_MAX_COLUMN_WIDTH, width + 2)
        if table.columns[index - 1].align == "right":
            for cell in sheet[letter]:
                cell.alignment = right

    # Freeze the header and switch on filters, so a 50,000-row export is
    # navigable rather than merely present.
    sheet.freeze_panes = "A2"
    if table.rows:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(table.columns))}{len(table.rows) + 1}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _as_number(value: str) -> float | int | str:
    """Coerce a formatted figure back to a number, or leave it alone."""
    if not value:
        return value
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        return value


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

_PAGE_W, _PAGE_H = 842.0, 595.0  # A4 landscape, in points — these tables are wide
_MARGIN = 32.0
_TITLE_SIZE = 15.0
_BODY_SIZE = 8.0
_HEADER_SIZE = 8.0
_LINE_H = 13.0
_ROWS_PER_PAGE = 34

#: Helvetica advance widths, in 1/1000 em, for every printable ASCII codepoint
#: from 32 to 126. Straight from the font's AFM metrics: text has to be measured
#: to be laid out, and a viewer will not do it for us.
# fmt: off
# Kept in rows of 16 rather than one number per line: this is a lookup table
# copied from the font's AFM metrics, and it is checked by counting entries.
_HELV = [
    278, 278, 355, 556, 556, 889, 667, 191, 333, 333, 389, 584, 278, 333, 278, 278,
    556, 556, 556, 556, 556, 556, 556, 556, 556, 556, 278, 278, 584, 584, 584, 556,
    1015, 667, 667, 722, 722, 667, 611, 778, 722, 278, 500, 667, 556, 833, 722, 778,
    667, 778, 722, 667, 611, 722, 667, 944, 667, 667, 611, 278, 278, 278, 469, 556,
    333, 556, 556, 500, 556, 556, 278, 556, 556, 222, 222, 500, 222, 833, 556, 556,
    556, 556, 333, 500, 278, 556, 500, 722, 500, 500, 500, 334, 260, 334, 584,
]
_HELV_BOLD = [
    278, 333, 474, 556, 556, 889, 722, 238, 333, 333, 389, 584, 278, 333, 278, 278,
    556, 556, 556, 556, 556, 556, 556, 556, 556, 556, 333, 333, 584, 584, 584, 611,
    975, 722, 722, 722, 722, 667, 611, 778, 722, 278, 556, 722, 611, 833, 722, 778,
    667, 778, 722, 667, 611, 722, 667, 944, 667, 667, 611, 333, 278, 333, 584, 556,
    333, 556, 611, 556, 611, 556, 333, 611, 611, 278, 278, 556, 278, 889, 611, 611,
    611, 611, 389, 556, 333, 611, 556, 778, 556, 556, 500, 389, 280, 389, 584,
]
# fmt: on


def _text_width(text: str, size: float, bold: bool = False) -> float:
    metrics = _HELV_BOLD if bold else _HELV
    total = 0
    for char in text:
        index = ord(char) - 32
        total += metrics[index] if 0 <= index < len(metrics) else 556
    return total * size / 1000.0


def _pdf_escape(text: str) -> str:
    r"""Escape for a PDF literal string, and drop what WinAnsi cannot show.

    The base-14 fonts are single-byte, so a rupee sign or a Devanagari product
    name has no glyph. Substituting rather than emitting a raw byte keeps the
    file valid; the column header already says the unit is rupees.
    """
    out = []
    for char in text:
        if char == "₹":
            out.append("Rs.")
        elif char in ("\\", "(", ")"):
            out.append("\\" + char)
        elif 32 <= ord(char) <= 126:
            out.append(char)
        elif ord(char) in (8211, 8212):  # en/em dash
            out.append("-")
        elif ord(char) in (8216, 8217):
            out.append("'")
        elif ord(char) in (8220, 8221):
            out.append('"')
        else:
            out.append("?")
    return "".join(out)


def _truncate(text: str, limit: float, size: float, bold: bool = False) -> str:
    """Clip to the column width, with an ellipsis, so cells never overlap."""
    if _text_width(text, size, bold) <= limit:
        return text
    ellipsis = ".."
    budget = limit - _text_width(ellipsis, size, bold)
    clipped = ""
    for char in text:
        if _text_width(clipped + char, size, bold) > budget:
            break
        clipped += char
    return clipped + ellipsis


def _column_widths(table: ReportTable) -> list[float]:
    """Share the page out in proportion to what each column actually holds."""
    usable = _PAGE_W - 2 * _MARGIN
    gutter = 8.0
    natural = []
    for index, column in enumerate(table.columns):
        widest = _text_width(column.header, _HEADER_SIZE, bold=True)
        # Sampled rather than exhaustive: measuring 50,000 rows to pick a column
        # width costs more than the layout is worth, and the first few hundred
        # are representative.
        for row in table.rows[:400]:
            if index < len(row):
                widest = max(widest, _text_width(row[index], _BODY_SIZE))
        natural.append(widest + gutter)

    total = sum(natural)
    if total <= usable:
        return natural
    # Over-wide: scale down proportionally, but keep every column readable.
    floor = 42.0
    scale = (usable - floor * len(natural)) / max(total - floor * len(natural), 1.0)
    return [floor + (width - floor) * scale for width in natural]


def to_pdf(table: ReportTable) -> bytes:
    """A paginated landscape table, one content stream per page."""
    widths = _column_widths(table)
    pages = [
        table.rows[start : start + _ROWS_PER_PAGE]
        for start in range(0, max(len(table.rows), 1), _ROWS_PER_PAGE)
    ] or [[]]
    stamp = datetime.now(UTC).strftime("%d %b %Y %H:%M UTC")

    streams = [
        _page_stream(table, widths, rows, number, len(pages), stamp)
        for number, rows in enumerate(pages, start=1)
    ]
    return _assemble(streams)


def _page_stream(
    table: ReportTable,
    widths: list[float],
    rows: Sequence[Sequence[str]],
    number: int,
    total: int,
    stamp: str,
) -> bytes:
    parts: list[str] = []

    def text(value: str, x: float, y: float, size: float, bold: bool = False) -> None:
        font = "/F2" if bold else "/F1"
        parts.append(
            f"BT {font} {size:.1f} Tf 1 0 0 1 {x:.2f} {y:.2f} Tm ({_pdf_escape(value)}) Tj ET"
        )

    def line(x1: float, y1: float, x2: float, y2: float, grey: float) -> None:
        parts.append(f"{grey:.2f} G 0.6 w {x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S")

    top = _PAGE_H - _MARGIN
    if number == 1:
        text(table.title, _MARGIN, top - 10, _TITLE_SIZE, bold=True)
        text(table.subtitle, _MARGIN, top - 26, 9.0)
        if table.truncated:
            text(
                f"Truncated to the first {len(table.rows)} rows.",
                _MARGIN,
                top - 39,
                8.0,
            )
        header_y = top - 56
    else:
        header_y = top - 12

    # Column headers, then a rule under them.
    x = _MARGIN
    for column, width in zip(table.columns, widths, strict=False):
        label = _truncate(column.header, width - 6, _HEADER_SIZE, bold=True)
        offset = (
            width - 6 - _text_width(label, _HEADER_SIZE, bold=True)
            if column.align == "right"
            else 0.0
        )
        text(label, x + 2 + offset, header_y, _HEADER_SIZE, bold=True)
        x += width
    line(_MARGIN, header_y - 4, _MARGIN + sum(widths), header_y - 4, 0.65)

    y = header_y - 4 - _LINE_H
    for row in rows:
        x = _MARGIN
        for index, (column, width) in enumerate(zip(table.columns, widths, strict=False)):
            value = row[index] if index < len(row) else ""
            clipped = _truncate(value, width - 6, _BODY_SIZE)
            offset = (
                width - 6 - _text_width(clipped, _BODY_SIZE) if column.align == "right" else 0.0
            )
            text(clipped, x + 2 + offset, y, _BODY_SIZE)
            x += width
        y -= _LINE_H

    if not rows and number == 1:
        text("No rows in this report.", _MARGIN, y, _BODY_SIZE)

    footer = f"StockSync Analytics  ·  generated {stamp}  ·  page {number} of {total}"
    text(_pdf_escape(footer), _MARGIN, _MARGIN - 10, 7.5)
    return "\n".join(parts).encode("latin-1", "replace")


def _assemble(streams: list[bytes]) -> bytes:
    """Objects, xref table and trailer — the file's plumbing.

    Object numbering: 1 catalogue, 2 page tree, 3 and 4 the two fonts, then a
    (page, content) pair per page. Byte offsets are recorded as each object is
    written, because the xref table has to point at them exactly.
    """
    count = len(streams)
    first_page_obj = 5
    page_ids = [first_page_obj + i * 2 for i in range(count)]

    objects: dict[int, bytes] = {
        1: b"<< /Type /Catalog /Pages 2 0 R >>",
        2: (
            "<< /Type /Pages /Count "
            + str(count)
            + " /Kids ["
            + " ".join(f"{pid} 0 R" for pid in page_ids)
            + "] >>"
        ).encode("ascii"),
        3: b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>",
        4: b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold "
        b"/Encoding /WinAnsiEncoding >>",
    }

    for index, stream in enumerate(streams):
        page_id = page_ids[index]
        content_id = page_id + 1
        objects[page_id] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {_PAGE_W:.0f} {_PAGE_H:.0f}] "
            f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> "
            f"/Contents {content_id} 0 R >>"
        ).encode("ascii")
        objects[content_id] = (
            f"<< /Length {len(stream)} >>\nstream\n".encode("ascii") + stream + b"\nendstream"
        )

    out = bytearray(b"%PDF-1.4\n")
    # A binary comment marks the file as binary for anything transferring it in
    # text mode, which would otherwise rewrite the line endings and corrupt it.
    out += b"%\xe2\xe3\xcf\xd3\n"

    offsets: dict[int, int] = {}
    for number in sorted(objects):
        offsets[number] = len(out)
        out += f"{number} 0 obj\n".encode("ascii") + objects[number] + b"\nendobj\n"

    xref_at = len(out)
    highest = max(objects) + 1
    out += f"xref\n0 {highest}\n".encode("ascii")
    out += b"0000000000 65535 f \n"
    for number in range(1, highest):
        out += f"{offsets[number]:010d} 00000 n \n".encode("ascii")
    out += (f"trailer\n<< /Size {highest} /Root 1 0 R >>\nstartxref\n{xref_at}\n%%EOF\n").encode(
        "ascii"
    )
    return bytes(out)
