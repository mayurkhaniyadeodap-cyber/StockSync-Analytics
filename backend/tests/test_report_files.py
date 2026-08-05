"""The three export writers.

The PDF is checked by parsing it the way a viewer does — resolve the catalogue
through the cross-reference table, walk to each page, read its content stream —
rather than by asserting on substrings. A PDF with a single wrong byte offset
still contains all the right substrings and still fails to open, so a check
that does not follow the offsets does not check the thing that breaks.
"""

from __future__ import annotations

import io
import re
import zipfile

import pytest
from openpyxl import load_workbook

from app.services.report_data import Column, ReportTable
from app.services.report_files import CONTENT_TYPES, render, to_csv, to_pdf, to_xlsx

COLUMNS = (
    Column("SKU"),
    Column("Product name"),
    Column("On hand", "right"),
    Column("Revenue (₹)", "right"),
    Column("Stock status"),
)


def table(count: int = 5, **overrides: object) -> ReportTable:
    body = [
        (f"DD-{i:04d}", f"Steel Bottle {i}", str(i * 10), f"{i * 99.5:.2f}", "In stock")
        for i in range(1, count + 1)
    ]
    defaults: dict[str, object] = {
        "title": "SKU performance report",
        "subtitle": f"{count} SKUs over the last 30 days",
        "columns": COLUMNS,
        "rows": body,
    }
    defaults.update(overrides)
    return ReportTable(**defaults)  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# a minimal PDF reader, used only by these tests
# ---------------------------------------------------------------------------


class Pdf:
    """Just enough of a PDF reader to prove the file is navigable."""

    def __init__(self, raw: bytes) -> None:
        self.raw = raw
        self.offsets = self._read_xref()

    def _read_xref(self) -> dict[int, int]:
        match = re.search(rb"startxref\s+(\d+)", self.raw)
        assert match, "no startxref"
        start = int(match.group(1))
        assert self.raw[start : start + 4] == b"xref", "startxref does not point at the table"

        header = re.match(rb"xref\s+(\d+)\s+(\d+)\s+", self.raw[start:])
        assert header, "malformed xref header"
        first, count = int(header.group(1)), int(header.group(2))
        body = self.raw[start + header.end() :]

        offsets: dict[int, int] = {}
        for index in range(count):
            entry = body[index * 20 : index * 20 + 20]
            assert len(entry) == 20, f"xref entry {index} is not 20 bytes"
            if entry[17:18] == b"n":
                offsets[first + index] = int(entry[:10])
        return offsets

    def object_at(self, number: int) -> bytes:
        """Follow the xref offset and read the object it claims is there."""
        offset = self.offsets[number]
        assert self.raw.startswith(f"{number} 0 obj".encode(), offset), (
            f"xref says object {number} is at {offset}, found {self.raw[offset : offset + 24]!r}"
        )
        end = self.raw.index(b"endobj", offset)
        return self.raw[offset:end]

    def catalog(self) -> bytes:
        root = re.search(rb"/Root (\d+) 0 R", self.raw)
        assert root, "trailer has no /Root"
        return self.object_at(int(root.group(1)))

    def pages(self) -> list[bytes]:
        pages_ref = re.search(rb"/Pages (\d+) 0 R", self.catalog())
        assert pages_ref, "catalogue has no /Pages"
        tree = self.object_at(int(pages_ref.group(1)))
        kids = re.search(rb"/Kids \[([^\]]*)\]", tree)
        assert kids, "page tree has no /Kids"
        numbers = [int(n) for n in re.findall(rb"(\d+) 0 R", kids.group(1))]

        count = re.search(rb"/Count (\d+)", tree)
        assert count and int(count.group(1)) == len(numbers), "/Count disagrees with /Kids"
        return [self.object_at(number) for number in numbers]

    def text_of(self, page: bytes) -> str:
        """The page's content stream, resolved through its /Contents reference."""
        ref = re.search(rb"/Contents (\d+) 0 R", page)
        assert ref, "page has no /Contents"
        obj = self.object_at(int(ref.group(1)))
        length = re.search(rb"/Length (\d+)", obj)
        assert length, "content stream has no /Length"
        start = obj.index(b"stream\n") + len(b"stream\n")
        stream = obj[start : start + int(length.group(1))]
        assert obj[start + int(length.group(1)) :].startswith(b"\nendstream"), "/Length is wrong"
        return " ".join(re.findall(r"\((.*?)\) Tj", stream.decode("latin-1")))


class TestPdf:
    def test_it_is_a_pdf(self) -> None:
        raw = to_pdf(table())

        assert raw.startswith(b"%PDF-1.4")
        assert raw.rstrip().endswith(b"%%EOF")

    def test_every_xref_offset_lands_on_its_object(self) -> None:
        """The one error that produces a file no viewer will open."""
        pdf = Pdf(to_pdf(table(120)))

        for number in pdf.offsets:
            pdf.object_at(number)  # asserts internally

    def test_the_catalogue_resolves_to_pages(self) -> None:
        pdf = Pdf(to_pdf(table()))

        assert b"/Type /Catalog" in pdf.catalog()
        assert len(pdf.pages()) == 1

    def test_long_reports_paginate(self) -> None:
        pdf = Pdf(to_pdf(table(200)))

        assert len(pdf.pages()) > 1

    def test_the_title_and_rows_are_in_the_content_stream(self) -> None:
        pdf = Pdf(to_pdf(table()))
        text = pdf.text_of(pdf.pages()[0])

        assert "SKU performance report" in text
        assert "DD-0001" in text
        assert "Steel Bottle 1" in text

    def test_every_page_is_numbered(self) -> None:
        pdf = Pdf(to_pdf(table(200)))
        pages = pdf.pages()

        for index, page in enumerate(pages, start=1):
            assert f"page {index} of {len(pages)}" in pdf.text_of(page)

    def test_a_glyphless_character_is_substituted_not_emitted(self) -> None:
        """Base-14 fonts are single-byte; a raw ₹ would corrupt the stream."""
        raw = to_pdf(table())

        assert "₹".encode() not in raw
        assert b"Rs." in raw

    def test_devanagari_becomes_a_placeholder_rather_than_breaking_the_file(self) -> None:
        pdf = Pdf(to_pdf(table(1, rows=[("DD-1", "साबुन दानी", "5", "10.00", "In stock")])))

        assert "DD-1" in pdf.text_of(pdf.pages()[0])

    def test_a_long_value_is_clipped_to_its_column(self) -> None:
        """Without clipping, a long product name runs under the next column."""
        long_name = "A" * 400
        pdf = Pdf(to_pdf(table(1, rows=[("DD-1", long_name, "5", "10.00", "In")])))
        text = pdf.text_of(pdf.pages()[0])

        assert long_name not in text
        assert ".." in text

    def test_parentheses_are_escaped(self) -> None:
        """An unescaped ( ends the string early and corrupts everything after it."""
        pdf = Pdf(to_pdf(table(1, rows=[("DD-1", "Bottle (750ml)", "5", "10.00", "In")])))

        assert r"Bottle \(750ml\)" in pdf.text_of(pdf.pages()[0])

    def test_an_empty_report_still_produces_a_valid_file(self) -> None:
        pdf = Pdf(to_pdf(table(0, rows=[], subtitle="0 SKUs")))

        assert len(pdf.pages()) == 1
        assert "No rows in this report." in pdf.text_of(pdf.pages()[0])

    def test_truncation_is_stated_on_the_page(self) -> None:
        pdf = Pdf(to_pdf(table(3, truncated=True)))

        assert "Truncated to the first 3 rows." in pdf.text_of(pdf.pages()[0])


class TestCsv:
    def test_excel_reads_the_encoding(self) -> None:
        """Without the BOM Excel uses the system codepage and mangles ₹ and é."""
        raw = to_csv(table())

        assert raw.startswith(b"\xef\xbb\xbf")
        assert raw.decode("utf-8-sig").startswith("SKU,Product name")

    def test_rfc4180_line_endings(self) -> None:
        assert b"\r\n" in to_csv(table())

    def test_non_ascii_survives(self) -> None:
        raw = to_csv(table(1, rows=[("DD-1", "Café Mug — ₹", "5", "10.00", "In")]))

        assert "Café Mug — ₹" in raw.decode("utf-8-sig")

    def test_a_comma_in_a_value_is_quoted(self) -> None:
        raw = to_csv(table(1, rows=[("DD-1", "Bottle, large", "5", "10.00", "In")]))

        assert '"Bottle, large"' in raw.decode("utf-8-sig")

    def test_headers_only_when_there_are_no_rows(self) -> None:
        text = to_csv(table(0, rows=[])).decode("utf-8-sig")

        assert text.strip() == "SKU,Product name,On hand,Revenue (₹),Stock status"


class TestXlsx:
    def test_it_is_a_real_workbook(self) -> None:
        raw = to_xlsx(table())

        assert "xl/workbook.xml" in zipfile.ZipFile(io.BytesIO(raw)).namelist()

    def test_figures_are_stored_as_numbers(self) -> None:
        """Stored as text they neither sum nor sort, and Excel flags every cell."""
        sheet = load_workbook(io.BytesIO(to_xlsx(table()))).active
        assert sheet is not None

        assert sheet["C2"].value == 10
        assert sheet["D2"].value == pytest.approx(99.5)
        assert isinstance(sheet["A2"].value, str)

    def test_the_header_stays_visible_and_filterable(self) -> None:
        sheet = load_workbook(io.BytesIO(to_xlsx(table()))).active
        assert sheet is not None

        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref == "A1:E6"

    def test_the_sheet_name_drops_characters_excel_rejects(self) -> None:
        sheet = load_workbook(io.BytesIO(to_xlsx(table(1, title="Sales: 2026/07 [draft]")))).active
        assert sheet is not None

        assert not set(sheet.title) & set(r"[]:*?/\\")

    def test_an_empty_report_has_a_header_but_no_filter(self) -> None:
        sheet = load_workbook(io.BytesIO(to_xlsx(table(0, rows=[])))).active
        assert sheet is not None

        assert sheet["A1"].value == "SKU"
        assert sheet.auto_filter.ref is None


class TestRender:
    @pytest.mark.parametrize("fmt", ["csv", "xlsx", "pdf"])
    def test_each_format_produces_bytes(self, fmt: str) -> None:
        assert render(table(), fmt)

    def test_an_unknown_format_is_refused(self) -> None:
        with pytest.raises(ValueError, match="unsupported report format"):
            render(table(), "docx")

    def test_every_format_has_a_content_type(self) -> None:
        assert set(CONTENT_TYPES) == {"csv", "xlsx", "pdf"}

    def test_the_three_formats_carry_the_same_values(self) -> None:
        """The point of one builder feeding three writers."""
        source = table(3)
        text = to_csv(source).decode("utf-8-sig")
        sheet = load_workbook(io.BytesIO(to_xlsx(source))).active
        assert sheet is not None
        pdf = Pdf(to_pdf(source))
        page = pdf.text_of(pdf.pages()[0])

        for row in range(1, 4):
            sku = f"DD-{row:04d}"
            assert sku in text
            assert sheet.cell(row=row + 1, column=1).value == sku
            assert sku in page


#: Payloads a spreadsheet would execute. The first is the DDE vector that
#: turns a product name into code execution in Excel.
DANGEROUS = [
    "=cmd|'/c calc'!A1",
    "=1+1",
    "+1+1",
    "-1+1",
    "@SUM(A1)",
    "\t=1+1",  # Excel strips leading whitespace before deciding
    "\r=1+1",
]


class TestFormulaInjection:
    """Product names and SKUs come from an uploaded sheet or from Shopify, so
    they are attacker-influenced. Exports exist to be emailed, which means the
    person a payload runs on is often not the person who imported it."""

    @pytest.mark.parametrize("payload", DANGEROUS)
    def test_csv_defuses_a_formula(self, payload: str) -> None:
        raw = to_csv(table(1, rows=[("DD-1", payload, "1", "1.00", "In")]))

        text = raw.decode("utf-8-sig")
        assert f"'{payload}" in text
        # The cell must not begin with the trigger character any more.
        cell = text.splitlines()[1].split(",")[1].lstrip('"')
        assert cell.startswith("'")

    @pytest.mark.parametrize("payload", DANGEROUS)
    def test_xlsx_never_writes_a_formula(self, payload: str) -> None:
        raw = to_xlsx(table(1, rows=[("DD-1", payload, "1", "1.00", "In")]))

        sheet = load_workbook(io.BytesIO(raw)).active
        assert sheet is not None
        assert sheet["B2"].data_type == "s"
        # The value is preserved — only its type is pinned. A carriage return
        # comes back as a newline because XML parsers are required to normalise
        # line endings (XML 1.0 §2.11); that is the format, not the guard.
        assert sheet["B2"].value == payload.replace("\r", "\n")

        xml = zipfile.ZipFile(io.BytesIO(raw)).read("xl/worksheets/sheet1.xml").decode()
        assert "<f>" not in xml

    def test_a_payload_in_the_sku_column_is_caught_too(self) -> None:
        """Not just product names — a SKU comes from the same sheet."""
        rows = [("=cmd|'/c calc'!A1", "Bottle", "1", "1.00", "In")]

        assert "'=cmd" in to_csv(table(1, rows=rows)).decode("utf-8-sig")
        sheet = load_workbook(io.BytesIO(to_xlsx(table(1, rows=rows)))).active
        assert sheet is not None
        assert sheet["A2"].data_type == "s"

    def test_xlsx_marks_the_cell_as_quoted_text(self) -> None:
        """quotePrefix is what makes Excel render it as text rather than re-parse."""
        raw = to_xlsx(table(1, rows=[("DD-1", "=1+1", "1", "1.00", "In")]))

        sheet = load_workbook(io.BytesIO(raw)).active
        assert sheet is not None
        assert sheet["B2"].quotePrefix is True

    @pytest.mark.parametrize(
        "safe",
        ["Bottle - 750ml", "Steel Bottle", "A+ Grade", "hello@example.com", "", "12.5"],
    )
    def test_ordinary_values_are_left_alone(self, safe: str) -> None:
        """A hyphen or @ inside a value is not a formula and must not be mangled."""
        text = to_csv(table(1, rows=[("DD-1", safe, "1", "1.00", "In")])).decode("utf-8-sig")

        assert "'" not in text.splitlines()[1]

    def test_the_pdf_needs_no_guard_but_still_shows_the_text(self) -> None:
        """A PDF cannot execute a cell; the value should survive readably."""
        pdf = Pdf(to_pdf(table(1, rows=[("DD-1", "=1+1", "1", "1.00", "In")])))

        assert "=1+1" in pdf.text_of(pdf.pages()[0])

    def test_numbers_are_still_numbers_after_the_guard(self) -> None:
        """The fix must not turn the numeric columns into text."""
        sheet = load_workbook(io.BytesIO(to_xlsx(table(3)))).active
        assert sheet is not None

        assert sheet["C2"].value == 10
        assert isinstance(sheet["D2"].value, float)


class TestXlsxStyling:
    def test_right_aligned_columns_keep_their_alignment(self) -> None:
        """Guards the shared-Alignment optimisation against a formatting regression."""
        sheet = load_workbook(io.BytesIO(to_xlsx(table(20)))).active
        assert sheet is not None

        for row in range(2, 22):
            assert sheet.cell(row=row, column=3).alignment.horizontal == "right"
            assert sheet.cell(row=row, column=1).alignment.horizontal is None

    def test_column_widths_cover_every_column(self) -> None:
        """The width scan moved into the append loop; right columns must not be skipped."""
        long_number = "1234567890123456"
        sheet = load_workbook(
            io.BytesIO(to_xlsx(table(1, rows=[("DD-1", "x", long_number, "1.00", "In")])))
        ).active
        assert sheet is not None

        assert sheet.column_dimensions["C"].width >= len(long_number)

    def test_one_alignment_object_is_shared(self) -> None:
        """The regression this guards cost ~2.5 s per 13,000-row export."""
        from unittest.mock import patch

        import app.services.report_files as module

        with patch.object(module, "Alignment", wraps=module.Alignment) as spy:
            to_xlsx(table(500))

        # One per right-aligned column at most, never one per cell.
        assert spy.call_count <= 4
